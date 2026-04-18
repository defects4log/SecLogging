import os
import subprocess
import json
from concurrent.futures import ThreadPoolExecutor, as_completed

api_key = "ollama"
api_url = "url"

def get_folder_name(use_description, use_explanation):
    if use_description and use_explanation:
        return "with_description_and_explanation"
    elif use_description:
        return "with_description"
    elif use_explanation:
        return "with_explanation"
    else:
        return "without_description_and_explanation"

def get_run_cmd(model_name, api_key, api_url, use_description, use_explanation):
    output_file = f"log_security/results/different_model/{model_name.replace(':', '_')}/{get_folder_name(use_description, use_explanation)}_result_test.csv"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
  
    batch_size = 1
    cmd = f"python analyze_with_llm.py --combined-csv log_security/results/combined_result_test.csv --overview-csv '../data/ground_Truth_test.csv' --output {output_file} --model {model_name}  --api-key {api_key} --api-url {api_url} --batch-size {batch_size} --temperature 1" + " --use-description" * use_description + " --use-explanation" * use_explanation
    return cmd, output_file

local_models = [
    "qwen2.5:72b-instruct",
    "llama3.3:latest",
    "qwen3:4b",
    "qwen3:32b",
    "deepseek-r1:32b"
]
local_api_url = "url"
local_api_key = "ollama"

online_models = [
    "deepseek-reasoner",
    "deepseek-chat"
]
online_api_url = "url"
online_api_key = "key"

def _run_one(cmd, output_file):

    
    proc = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    log_path = output_file + ".log"
    try:
        with open(log_path, "w", encoding="utf-8") as lf:
            if proc.stdout:
                lf.write(proc.stdout)
            if proc.stderr:
                lf.write("\n[stderr]\n")
                lf.write(proc.stderr)
    except Exception:
        pass
    return {
        'cmd': cmd,
        'output_file': output_file,
        'status': 'ok' if proc.returncode == 0 else 'failed',
        'returncode': proc.returncode,
        'log': log_path
    }

def _run_parallel(cmds, output_files, max_parallel):
    """Run many commands in parallel with a max concurrency limit."""
    results = []
    if not cmds:
        return results
    workers = max(1, min(max_parallel, len(cmds)))
    print(f"Running {len(cmds)} tasks with max_parallel={workers} ...")
    with ThreadPoolExecutor(max_workers=workers) as ex:
        future_map = {ex.submit(_run_one, c, o): (c, o) for c, o in zip(cmds, output_files)}
        for fut in as_completed(future_map):
            res = fut.result()
            print(f"[{res['status']}] rc={res['returncode']} -> {res['output_file']}")
            results.append(res)
  
    succeeded = sum(1 for r in results if r['status'] == 'ok')
    failed = sum(1 for r in results if r['status'] == 'failed')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    print(f"Done. ok={succeeded}, failed={failed}, skipped={skipped}, total={len(results)}")
    return results


def run_local_model_parallel(model_names, api_key, api_url, max_parallel=5):
    cmds = []
    output_files = []

    for use_description in [True, False]:
        for use_explanation in [True, False]:
            for model_name in model_names:
                cmd, output_file = get_run_cmd(model_name, api_key, api_url, use_description, use_explanation)
                cmds.append(cmd)
                output_files.append(output_file)
    return _run_parallel(cmds, output_files, max_parallel)

def run_online_model_parallel(model_names, api_key, api_url, max_parallel=5):
    cmds = []
    output_files = []
    for model_name in model_names:
        for use_description in [True, False]:
            for use_explanation in [True, False]:
                cmd, output_file = get_run_cmd(model_name, api_key, api_url, use_description, use_explanation)
                cmds.append(cmd)
                output_files.append(output_file)
    return _run_parallel(cmds, output_files, max_parallel)

from calculate_accuracy import evaluate
ground_truth_file = "log_security/results/Datas/ground_Truth.csv"

model_names = local_models + online_models

def evaluate_all_models(model_names, ground_truth_file):
    results = []
    for model_name in model_names:
        for use_description in [True, False]:
            for use_explanation in [True, False]:
                folder_name = get_folder_name(use_description, use_explanation)
                result_file = f"log_security/results/different_model/{model_name.replace(':', '_')}/merged/{folder_name}_merged_result.csv"
                if os.path.exists(result_file):
                    category_accuracy, pattern_accuracy, results_by_category, results_by_issue = evaluate(result_file, ground_truth_file)
                    res = {}
                    res['model_name'] = model_name
                    res['use_description'] = use_description
                    res['use_explanation'] = use_explanation
                    res['result_file'] = result_file
                    res['category_accuracy'] = category_accuracy
                    res['pattern_accuracy'] = pattern_accuracy
                    res['results_by_category'] = results_by_category
                    results.append(res)
                else:
                    print(f"Warning: result file {result_file} does not exist, skipping evaluation.")
    return results

def to_excel_table(evaluation_results, output_excel_file):
    import os
    from collections import defaultdict
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError as e:
        raise RuntimeError("pip install openpyxl") from e

    by_model = defaultdict(dict)
    for item in evaluation_results or []:
        key = (bool(item.get("use_description")), bool(item.get("use_explanation")))
        by_model[item["model_name"]][key] = item

    columns = [
        ((True, True), "D+E"),
        ((True, False), "D"),
        ((False, True), "E"),
        ((False, False), "None"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    model_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(style="thin", color="FFAAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions["A"].width = 34
    for i, _ in enumerate(columns, start=2):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = 18

    row = 1
    for model_name, combos in by_model.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(columns))
        cell = ws.cell(row=row, column=1, value=model_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center
        cell.fill = model_fill
        cell.border = border

        for c in range(2, 2 + len(columns) - 1 + 1):
            cc = ws.cell(row=row, column=c)
            cc.fill = model_fill
            cc.border = border
        row += 1

        ws.cell(row=row, column=1, value="Metric").font = bold
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = border
        for idx, (_, col_title) in enumerate(columns, start=2):
            h = ws.cell(row=row, column=idx, value=col_title)
            h.font = bold
            h.alignment = center
            h.fill = header_fill
            h.border = border
        row += 1

        def write_metric_row(metric_name, extractor, number_format=None):
            nonlocal row
            ws.cell(row=row, column=1, value=metric_name).alignment = left
            ws.cell(row=row, column=1).border = border
            for idx, (key, _) in enumerate(columns, start=2):
                val = extractor(combos.get(key))
                cell = ws.cell(row=row, column=idx, value=val)
                if isinstance(val, (int, float)) and number_format:
                    cell.number_format = number_format
                cell.alignment = center
                cell.border = border
            row += 1

        write_metric_row(
            "Overall Category Accuracy",
            lambda item: None if not item else item.get("category_accuracy"),
            number_format="0.00%",
        )
        write_metric_row(
            "Overall Pattern Accuracy",
            lambda item: None if not item else item.get("pattern_accuracy"),
            number_format="0.00%",
        )

        row += 1

        all_categories = set()
        for item in combos.values():
            cats = (item or {}).get("results_by_category") or {}
            all_categories.update(cats.keys())
        for cat in sorted(all_categories):
            write_metric_row(
                f"{cat} - Count",
                lambda item, c=cat: None if not item else ((item.get("results_by_category") or {}).get(c, {}).get("count")),
                number_format="0",
            )
            write_metric_row(
                f"{cat} - Category Accuracy",
                lambda item, c=cat: None if not item else ((item.get("results_by_category") or {}).get(c, {}).get("category_accuracy")),
                number_format="0.00%",
            )
            write_metric_row(
                f"{cat} - Pattern Accuracy",
                lambda item, c=cat: None if not item else ((item.get("results_by_category") or {}).get(c, {}).get("pattern_accuracy")),
                number_format="0.00%",
            )
        row += 1

    out_dir = os.path.dirname(output_excel_file)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_excel_file)


def build_merge_and_compare_cmd(llm_res_path, combined_path, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    cmd = f"python merge_and_compare_similarity.py --llm-results {llm_res_path} --combined-results {combined_path} --output {output_path}"
    return cmd

combined_path = "log_security/results/Datas/combined_result.csv"
def run_merge_and_compare():
    model_names = local_models + online_models

    for model_name in model_names:
        for use_description in [True, False]:
            for use_explanation in [True, False]:
                folder_name = get_folder_name(use_description, use_explanation)
                llm_res_path = f"log_security/results/different_model/{model_name.replace(':', '_')}/merged/{folder_name}_merged_result.csv"

                output_path = f"log_security/results/different_model/{model_name.replace(':', '_')}/merged/{folder_name}__merged_similarity_analysis_results.csv"
                if os.path.exists(llm_res_path) and os.path.exists(combined_path):
                    cmd = build_merge_and_compare_cmd(llm_res_path, combined_path, output_path)
                    print(f"Running merge and compare for {model_name} ({folder_name}) ...")
                 
                    subprocess.run(cmd, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                else:
                    print(f"Skipping {model_name} ({folder_name}) due to missing files.")

import pandas as pd

def merge_res(res_path, test_res_path, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if not os.path.exists(res_path):
        print(f"Warning: {res_path} does not exist, skipping merge.")
        return
    if not os.path.exists(test_res_path):
        print(f"Warning: {test_res_path} does not exist, skipping merge.")
        return
    df1 = pd.read_csv(res_path)
    df2 = pd.read_csv(test_res_path)
    
    common_cols = ['url', 'filename', 'issue_title', 'description', 'category', 'pattern', 'problem', 'fix_recommendation', 'fixed_code', 'analysis_status']
    df1_common = df1[common_cols]
    df2_common = df2[common_cols]

    merged_df = pd.concat([df1_common, df2_common], ignore_index=True)

    merged_df.to_csv(output_path, index=False)
    print(f"Merged results saved to {output_path}")

def merge_all_res():
    model_names = online_models + local_models

    for model_name in model_names:
        for use_description in [True, False]:
            for use_explanation in [True, False]:
                folder_name = get_folder_name(use_description, use_explanation)
                res_path = f"log_security/results/different_model/{model_name.replace(':', '_')}/{folder_name}_result.csv"
                test_res_path = f"log_security/results/different_model/{model_name.replace(':', '_')}/{folder_name}_result_test.csv"
                output_path = f"log_security/results/different_model/{model_name.replace(':', '_')}/{folder_name}_merged_result.csv"
                merge_res(res_path, test_res_path, output_path)



run_merge_and_compare()

